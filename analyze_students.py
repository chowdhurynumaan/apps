import openpyxl
from collections import defaultdict

# Load the Excel file
wb = openpyxl.load_workbook('Copy of studentdata.xlsm')
ws = wb['studentdata']

# Get headers
headers = {i: cell.value for i, cell in enumerate([cell for cell in ws[1]])}

# Extract column indices
name_idx = None
gender_idx = None
weekend_idx = None
evening_idx = None
grade_idx = None
group_idx = None
level_idx = None

for i, header in headers.items():
    if header == 'Name': name_idx = i
    elif header == 'Gender': gender_idx = i
    elif header == 'Weekend 2025': weekend_idx = i
    elif header == 'Evening 2025': evening_idx = i
    elif header == 'Grade': grade_idx = i
    elif header == 'Group': group_idx = i
    elif header == 'Level': level_idx = i

# Parse data
students = []
for row in ws.iter_rows(min_row=2, values_only=False):
    row_data = {i: cell.value for i, cell in enumerate(row)}
    if row_data.get(name_idx):  # Only include rows with names
        students.append(row_data)

total = len(students)

print("=" * 80)
print("STUDENT ROSTER STATISTICS")
print("=" * 80)
print(f"\n📊 TOTAL STUDENTS: {total}")

# Gender
print(f"\n👥 GENDER BREAKDOWN:")
girls = sum(1 for s in students if s.get(gender_idx) == 'F')
boys = sum(1 for s in students if s.get(gender_idx) == 'M')
other = total - girls - boys
print(f"  Girls (F): {girls}")
print(f"  Boys (M): {boys}")
if other > 0:
    print(f"  Other/Unknown: {other}")

# Weekend enrollment
weekend_students = [s for s in students if s.get(weekend_idx)]
print(f"\n📅 WEEKEND 2025 ENROLLMENT: {len(weekend_students)} students")

weekend_classes = defaultdict(int)
weekend_girls = defaultdict(int)
weekend_boys = defaultdict(int)
for s in weekend_students:
    cls = s.get(weekend_idx)
    if cls:
        weekend_classes[cls] += 1
        if s.get(gender_idx) == 'F':
            weekend_girls[cls] += 1
        elif s.get(gender_idx) == 'M':
            weekend_boys[cls] += 1

print(f"  By class:")
for cls in sorted(weekend_classes.keys(), key=lambda x: (str(x).rstrip('abcABC'), str(x))):
    print(f"    Class {cls}: {weekend_classes[cls]} ({weekend_girls.get(cls, 0)}G, {weekend_boys.get(cls, 0)}B)")

print(f"  By gender:")
weekend_g = sum(1 for s in weekend_students if s.get(gender_idx) == 'F')
weekend_b = sum(1 for s in weekend_students if s.get(gender_idx) == 'M')
print(f"    Girls: {weekend_g}")
print(f"    Boys: {weekend_b}")

# Evening enrollment
evening_students = [s for s in students if s.get(evening_idx)]
print(f"\n🌙 EVENING 2025 ENROLLMENT: {len(evening_students)} students")

evening_classes = defaultdict(int)
evening_girls = defaultdict(int)
evening_boys = defaultdict(int)
for s in evening_students:
    cls = s.get(evening_idx)
    if cls:
        evening_classes[cls] += 1
        if s.get(gender_idx) == 'F':
            evening_girls[cls] += 1
        elif s.get(gender_idx) == 'M':
            evening_boys[cls] += 1

print(f"  By class:")
for cls in sorted(evening_classes.keys(), key=lambda x: (str(x).rstrip('abcABC'), str(x))):
    print(f"    Class {cls}: {evening_classes[cls]} ({evening_girls.get(cls, 0)}G, {evening_boys.get(cls, 0)}B)")

print(f"  By gender:")
evening_g = sum(1 for s in evening_students if s.get(gender_idx) == 'F')
evening_b = sum(1 for s in evening_students if s.get(gender_idx) == 'M')
print(f"    Girls: {evening_g}")
print(f"    Boys: {evening_b}")

# Grades
print(f"\n🎓 GRADE LEVELS:")
grades = defaultdict(int)
for s in students:
    grade = s.get(grade_idx)
    if grade:
        grades[grade] += 1
    else:
        grades['(Not specified)'] += 1
for grade in sorted(grades.keys()):
    print(f"  {grade}: {grades[grade]}")

# Groups
print(f"\n📖 GROUPS (Hifz Programs):")
groups = defaultdict(int)
for s in students:
    group = s.get(group_idx)
    if group:
        groups[str(group)] += 1
    else:
        groups['(No group)'] += 1
for group in sorted(groups.keys(), key=str):
    print(f"  {group}: {groups[group]}")

# Levels
print(f"\n📚 LEVELS:")
levels = defaultdict(int)
for s in students:
    level = s.get(level_idx)
    if level:
        levels[str(level)] += 1
    else:
        levels['(Not specified)'] += 1
for level in sorted(levels.keys(), key=str):
    print(f"  {level}: {levels[level]}")

# Enrollment breakdown
both = sum(1 for s in students if s.get(weekend_idx) and s.get(evening_idx))
only_w = sum(1 for s in students if s.get(weekend_idx) and not s.get(evening_idx))
only_e = sum(1 for s in students if not s.get(weekend_idx) and s.get(evening_idx))
neither = sum(1 for s in students if not s.get(weekend_idx) and not s.get(evening_idx))

print(f"\n🔄 ENROLLMENT BREAKDOWN:")
print(f"  Both Weekend and Evening: {both}")
print(f"  Weekend only: {only_w}")
print(f"  Evening only: {only_e}")
print(f"  Not enrolled: {neither}")

print("\n" + "=" * 80)
